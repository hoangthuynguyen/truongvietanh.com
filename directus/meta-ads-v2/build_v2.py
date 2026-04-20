#!/usr/bin/env python3
"""
Lion Camp 2026 — Meta Ads v2 Builder
Generates the full v2 system:
- 24 Campaigns (Region × Level × Funnel)
- 96 Ad Sets (4 per campaign: Broad / Interest / Lookalike / Retarget)
- 30 Ads (10 each for TH / THCS / THPT, mapped across funnels)
- Pixel Events, Audiences, Budget pacing, KPI targets, Checklist

Outputs: MetaAds_LionCamp_2026_v2.xlsx + CSVs + JSON for Apps Script.
"""
import csv
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.dirname(os.path.abspath(__file__))

# =========================================================
# CORE CONFIG
# =========================================================
TOTAL_BUDGET = 420_000_000   # VND, 60 days
DAYS = 60
DAILY_TOTAL = TOTAL_BUDGET // DAYS  # 7,000,000 VND/day

# Funnel split (% of daily)
FUNNEL_PCT = {"Quiz": 0.50, "VSL": 0.30, "Sales": 0.20}
# Level split (% of daily across all regions per funnel)
LEVEL_PCT  = {"THPT": 0.35, "THCS": 0.30, "TH": 0.35}

# Region weights inside each level
REGION_WEIGHT = {
    "THPT": {"GV": 0.50, "BT": 0.50},
    "THCS": {"GV": 0.50, "BT": 0.50},
    "TH":   {"GV": 0.30, "BT": 0.30, "CL": 0.20, "RG": 0.20},
}

REGION_LABEL = {"GV": "Gò Vấp", "BT": "Bình Tân", "CL": "Cần Giuộc", "RG": "Rạch Giá"}
LEVEL_LABEL  = {"TH": "Tiểu học", "THCS": "THCS", "THPT": "THPT"}

SLOTS = {
    ("GV", "TH"): 50,  ("GV", "THCS"): 100, ("GV", "THPT"): 100,
    ("BT", "TH"): 50,  ("BT", "THCS"): 100, ("BT", "THPT"): 100,
    ("CL", "TH"): 50,
    ("RG", "TH"): 50,
}

# =========================================================
# 1) CAMPAIGN MATRIX (24 campaigns)
# =========================================================
def build_campaigns():
    rows = []
    for level, region_w in REGION_WEIGHT.items():
        for region, rw in region_w.items():
            for funnel, fp in FUNNEL_PCT.items():
                daily = round(DAILY_TOTAL * fp * LEVEL_PCT[level] * rw / 1000) * 1000  # round to 1k VND
                total = daily * DAYS
                name = f"{region}_{level}_{funnel}_LionCamp_2026"
                slot = SLOTS.get((region, level), 0)
                objective = {
                    "Quiz":  "Leads (Conversion: QuizComplete)",
                    "VSL":   "Conversions (ThruPlay → ViewLP)",
                    "Sales": "Conversions (Lead/Booking)",
                }[funnel]
                rows.append({
                    "Campaign Name": name,
                    "Region": REGION_LABEL[region],
                    "Level": LEVEL_LABEL[level],
                    "Funnel": funnel,
                    "Slot Goal": slot,
                    "Objective": objective,
                    "CBO": "Yes (ABO for Test, CBO sau Day 7)",
                    "Daily Budget (VND)": daily,
                    "Total Budget (VND)": total,
                    "Bid Strategy": "Lowest cost (cap khi scale)",
                    "Start": "2026-04-18",
                    "End": "2026-06-15",
                    "Special Ad Category": "None",
                })
    return rows

# =========================================================
# 2) AD SETS (96 = 24 campaigns × 4 adsets)
# =========================================================
ADSET_TYPES = [
    {"suffix": "Broad",     "audience": "Broad AI (no interest)",            "share": 0.30},
    {"suffix": "Interest",  "audience": "Interest: Parenting + Education",   "share": 0.25},
    {"suffix": "LAL",       "audience": "Lookalike 1-2% (LeadDB + Quiz)",    "share": 0.20},
    {"suffix": "RMK",       "audience": "Retarget: Web 30d + Video 50% + Form 14d", "share": 0.25},
]

INTEREST_BY_LEVEL = {
    "TH":   "Tiểu học, Trại hè, Phụ huynh, Tiếng Anh trẻ em, Kỹ năng sống",
    "THCS": "Cha mẹ tuổi teen, Học thêm, IELTS Junior, Định hướng nghề",
    "THPT": "Du học, Hướng nghiệp, Đại học, IELTS, Kỹ năng mềm",
}

LOCATION_BY_REGION = {
    "GV": "Gò Vấp + Bán kính 5km + Q12, Tân Bình, Bình Thạnh",
    "BT": "Bình Tân + Bán kính 5km + Q6, Tân Phú, Bình Chánh",
    "CL": "Cần Giuộc, Long An (toàn huyện) + Bán kính 8km",
    "RG": "Rạch Giá, Kiên Giang (toàn TP) + Bán kính 8km",
}

AGE_BY_LEVEL = {"TH": "28-42", "THCS": "32-50", "THPT": "38-55"}

def build_adsets(campaigns):
    rows = []
    for c in campaigns:
        # parse code from name
        region_code = c["Campaign Name"].split("_")[0]
        level_code = c["Campaign Name"].split("_")[1]
        funnel = c["Funnel"]
        for at in ADSET_TYPES:
            adset_name = f"{c['Campaign Name'].replace('_LionCamp_2026','')}_{at['suffix']}"
            daily = round(c["Daily Budget (VND)"] * at["share"] / 1000) * 1000
            audience = at["audience"]
            if at["suffix"] == "Interest":
                audience += " | " + INTEREST_BY_LEVEL[level_code]
            placement = "Auto (Reels + Feed + Stories)" if funnel != "Sales" else "Manual: FB Feed + IG Feed + Reels"
            optimization = {
                "Quiz":  "Conversion (QuizComplete) — 7d click",
                "VSL":   "ThruPlay → Custom: ViewLP60s",
                "Sales": "Conversion (Lead) — 7d click + 1d view",
            }[funnel]
            rows.append({
                "Campaign": c["Campaign Name"],
                "Ad Set Name": adset_name,
                "Audience": audience,
                "Exclude": "Existing Leads 90d + Customers 365d",
                "Locations": LOCATION_BY_REGION[region_code],
                "Age": AGE_BY_LEVEL[level_code],
                "Gender": "All",
                "Placement": placement,
                "Daily Budget (VND)": daily,
                "Optimization": optimization,
                "Attribution": "7-day click, 1-day view",
                "Pacing": "Standard",
            })
    return rows

# =========================================================
# 3) ADS (30 ads — 10 per level)
# =========================================================
ADS_BY_LEVEL = {
    "TH": [
        # (Theme, Hook, Headline, CTA, Funnel)
        ("Pain",       "Mùa hè này con bạn sẽ làm gì? Ở nhà với điện thoại?", "Hè Vàng — Đừng để lãng phí",         "Đăng ký ngay",  "Quiz"),
        ("Outcome",    "4 tuần — con tự tin giao tiếp tiếng Anh cơ bản",      "Trại hè Lion Camp Tiểu học",         "Tìm hiểu thêm","VSL"),
        ("Social",     "“Sau trại hè, con tự tin hơn hẳn” — PH Gò Vấp",       "1.200+ phụ huynh đã chọn Lion",      "Xem testimonial","VSL"),
        ("Urgency",    "Còn 18 suất Tiểu học Gò Vấp — Khai giảng 15/06",      "Hết sớm hơn dự kiến!",                "Giữ chỗ ngay",  "Sales"),
        ("Activity",   "Học qua dự án + thể thao + kỹ năng sống",             "Lịch trại hè 2026",                   "Xem chi tiết",  "Quiz"),
        ("PainLocal",  "PH Bình Tân: Hè con ở nhà cả ngày với TV?",           "Trại hè ngay tại Bình Tân",           "Đăng ký",       "Quiz"),
        ("OfferLocal", "Trại hè Tiểu học gần nhà — Cần Giuộc + Rạch Giá",     "Không cần đi xa",                     "Tư vấn miễn phí","Sales"),
        ("ProofVideo", "Xem 1 ngày của con tại Lion Camp",                    "Ngày của Lion (video 60s)",           "Xem video",     "VSL"),
        ("Bonus",      "Đăng ký sớm — Tặng bộ sách Tiếng Anh + áo trại",      "Ưu đãi Early Bird",                   "Nhận ngay",     "Sales"),
        ("LastCall",   "Còn 7 ngày để giữ suất 50/lớp Tiểu học",              "Đóng đăng ký 30/04",                  "Đăng ký gấp",   "Sales"),
    ],
    "THCS": [
        ("Pain",       "Con bắt đầu mất định hướng, mê game, lười học?",      "THCS — Tuổi vàng định hướng",         "Test miễn phí","Quiz"),
        ("Identity",   "Độ tuổi 11–15 quyết định 80% tương lai con",          "Lion Camp THCS",                      "Tìm hiểu",     "VSL"),
        ("Peer",       "Bạn bè con đang đi trước — Còn con bạn?",             "Đừng để con bị bỏ lại",               "Đăng ký quiz", "Quiz"),
        ("Skill",      "Tư duy + Tiếng Anh + Kỹ năng — All in 1 trại hè",     "Combo phát triển toàn diện",          "Xem chương trình","VSL"),
        ("Urgency",    "Đã full 70% suất THCS Gò Vấp — Còn 30 suất",          "Sắp đóng đăng ký",                    "Giữ chỗ",      "Sales"),
        ("Outcome",    "Sau 4 tuần: Con biết mình muốn gì + lập kế hoạch",    "Outcome thực tế",                     "Xem case",     "VSL"),
        ("Schedule",   "Lịch trại 4 tuần — Cuối tuần được về nhà",            "Lịch THCS 2026",                      "Xem lịch",     "Quiz"),
        ("ParentVoice","“Con tôi thay đổi 180°” — PH chia sẻ",                "Câu chuyện thật",                     "Đọc thêm",     "VSL"),
        ("Local",      "Trại hè THCS ngay Bình Tân — Đi học gần nhà",         "Tiện cho phụ huynh đi làm",           "Đăng ký",      "Sales"),
        ("LastCall",   "Đợt 1 hết — Đợt 2 chỉ còn 12 suất",                   "Đợt cuối",                            "Đăng ký gấp",  "Sales"),
    ],
    "THPT": [
        ("Pain",       "Con bạn đã biết mình muốn học gì sau lớp 12 chưa?",   "THPT — Cuộc đua bắt đầu",             "Test định hướng","Quiz"),
        ("Career",     "Trại hè định hướng nghề nghiệp + soft skill",         "Career Lab Lion Camp",                "Tìm hiểu",     "VSL"),
        ("StudyAbroad","Chuẩn bị du học từ THPT — Roadmap 3 năm",             "Du học từ A-Z",                       "Xem lộ trình", "VSL"),
        ("Proof",      "Học sinh thật — Đỗ NUS, NTU, RMIT",                   "Alumni Lion Camp",                    "Xem hồ sơ",    "VSL"),
        ("Urgency",    "Đóng đăng ký THPT sau 7 ngày — Còn 22 suất",          "Khai giảng 15/06",                    "Đăng ký gấp",  "Sales"),
        ("IELTS",      "IELTS bridge + Soft skill trong 4 tuần",              "Combo IELTS + Hướng nghiệp",          "Xem khóa",     "Quiz"),
        ("Network",    "Networking với mentor + alumni đại học top",          "Mentor 1:1",                          "Đặt lịch",     "Sales"),
        ("ParentROI",  "Đầu tư 1 hè = Tiết kiệm 2 năm sai hướng",             "ROI rõ ràng",                         "Tư vấn",       "Sales"),
        ("Local",      "THPT Bình Tân — Trại hè ngay tại quận",               "Khỏi đưa đón xa",                     "Đăng ký",      "Quiz"),
        ("LastCall",   "Còn đúng 14 suất THPT — Đóng đăng ký 25/05",          "Final Call",                          "Giữ chỗ",      "Sales"),
    ],
}

def build_ads(adsets):
    rows = []
    for level, ads in ADS_BY_LEVEL.items():
        for i, (theme, hook, headline, cta, ad_funnel) in enumerate(ads, start=1):
            ad_name = f"{level}_AD{i:02d}_{theme}"
            # match all adsets where adset belongs to a campaign of this level + this funnel
            for ads_row in adsets:
                cmp_name = ads_row["Campaign"]
                _region, _level, _funnel = cmp_name.split("_")[:3]
                if _level == level and _funnel == ad_funnel:
                    landing = {
                        "Quiz":  f"https://truongvietanh.com/lioncamp/quiz?utm_source=fb&utm_campaign={cmp_name}",
                        "VSL":   f"https://truongvietanh.com/lioncamp/vsl?utm_source=fb&utm_campaign={cmp_name}",
                        "Sales": f"https://truongvietanh.com/lioncamp/sales?utm_source=fb&utm_campaign={cmp_name}",
                    }[ad_funnel]
                    rows.append({
                        "Campaign": cmp_name,
                        "Ad Set": ads_row["Ad Set Name"],
                        "Ad Name": f"{cmp_name.replace('_LionCamp_2026','')}_{ad_name}",
                        "Format": "Single Video 9:16 + Square 1:1 (Dynamic)",
                        "Theme": theme,
                        "Primary Text Hook": hook,
                        "Headline": headline,
                        "CTA Button": cta,
                        "Landing URL": landing,
                    })
    return rows

# =========================================================
# 4) PIXEL EVENTS (8 events)
# =========================================================
PIXEL_EVENTS = [
    (1, "PageView",          "Tất cả LP",                            None,        "Standard"),
    (2, "ViewContent",       "Vào Sales/VSL/Quiz LP",                 0,          "Standard"),
    (3, "Lead_QuizStart",    "Bắt đầu quiz",                          50_000,     "Custom"),
    (4, "Lead_QuizComplete", "Submit quiz + email/sđt",               150_000,    "Custom"),
    (5, "Lead_VSL_View60s",  "Xem VSL ≥ 60s",                         100_000,    "Custom"),
    (6, "Lead_FormSubmit",   "Submit form Sales (Đặt lịch tư vấn)",   500_000,    "Custom (Lead)"),
    (7, "Booking_Confirmed", "Sales xác nhận lịch tư vấn",            1_500_000,  "Offline/CAPI"),
    (8, "Enroll_Paid",       "Phụ huynh đặt cọc/đóng học phí",        12_000_000, "Offline/CAPI"),
]

# =========================================================
# 5) AUDIENCES (≥15)
# =========================================================
AUDIENCES = [
    ("Custom", "CA_LeadDB_All",          "Lead Database 12 tháng (CSV upload)",    "30k+", "180d", "LAL source + Exclude"),
    ("Custom", "CA_QuizComplete_90d",    "Pixel: Lead_QuizComplete 90d",            "5-10k","90d",  "RMK + LAL source"),
    ("Custom", "CA_QuizStart_NoSubmit",  "QuizStart - QuizComplete (90d)",          "3-5k", "90d",  "RMK Quiz funnel"),
    ("Custom", "CA_VSL_50pct_30d",       "Video: 50%+ trong 30d",                   "10-30k","30d", "RMK Warm"),
    ("Custom", "CA_VSL_75pct_30d",       "Video: 75%+ trong 30d",                   "5-15k","30d", "RMK Hot"),
    ("Custom", "CA_LP_Sales_30d",        "Web visitors Sales LP 30d",               "8-20k","30d", "RMK Sales"),
    ("Custom", "CA_LP_VSL_30d",          "Web visitors VSL LP 30d",                 "8-25k","30d", "RMK VSL"),
    ("Custom", "CA_FormSubmit_NoBooking","FormSubmit - Booking_Confirmed 14d",      "1-3k", "14d", "RMK LastCall"),
    ("Custom", "CA_FB_Engagers_90d",     "Page engagement FB 90d",                  "20-40k","90d","TOFU Warm"),
    ("Custom", "CA_IG_Engagers_90d",     "IG engagement 90d",                       "15-30k","90d","TOFU Warm"),
    ("LAL",    "LAL_1pct_LeadDB",        "Lookalike 1% từ CA_LeadDB_All — VN",      "300k", "Open","TOFU Cold"),
    ("LAL",    "LAL_1pct_QuizComplete",  "Lookalike 1% từ Quiz — VN",               "300k", "Open","TOFU Cold"),
    ("LAL",    "LAL_1pct_THPT_GVBT",     "LAL 1% từ Lead THPT GV+BT",               "300k", "Open","THPT campaigns"),
    ("LAL",    "LAL_2pct_TH_GVBT",       "LAL 2% từ Lead TH GV+BT",                 "600k", "Open","TH campaigns"),
    ("Saved",  "SA_Parents_28_55_VN",    "Parents 28-55 + Education + Parenting",   "2-5M", "Open","Broad/Interest"),
    ("Saved",  "SA_GoVap_Bantanrik5km",  "Bán kính 5km Gò Vấp + interest education","200-400k","Open","Local GV"),
    ("Saved",  "SA_BinhTan_Local",       "Bán kính 5km Bình Tân + Parents",         "200-400k","Open","Local BT"),
]

# =========================================================
# 6) BUDGET PACING
# =========================================================
def build_pacing(campaigns):
    # 60 ngày chia thành 5 tuần x 12 ngày? Thực tế 8 tuần (~7 ngày)
    # Phân bổ theo PDF gốc nhưng cho v2: scale Quiz đầu, dồn Sales/RMK cuối.
    weeks = [
        ("Tuần 1",   "18/04 → 24/04", 0.55, 0.30, 0.10, 0.05, "Scale Quiz, test creative"),
        ("Tuần 2",   "25/04 → 01/05", 0.50, 0.30, 0.15, 0.05, "Scale Quiz, bắt đầu VSL"),
        ("Tuần 3",   "02/05 → 08/05", 0.45, 0.30, 0.20, 0.05, "Scale VSL warm"),
        ("Tuần 4",   "09/05 → 15/05", 0.40, 0.30, 0.20, 0.10, "Push VSL + Sales"),
        ("Tuần 5",   "16/05 → 22/05", 0.30, 0.30, 0.25, 0.15, "Tăng Sales + RMK"),
        ("Tuần 6",   "23/05 → 29/05", 0.20, 0.25, 0.30, 0.25, "Sales push mạnh"),
        ("Tuần 7",   "30/05 → 05/06", 0.15, 0.20, 0.35, 0.30, "Last Call + RMK heavy"),
        ("Tuần 8",   "06/06 → 15/06", 0.10, 0.15, 0.40, 0.35, "ALL-IN closing 10 ngày cuối"),
    ]
    out = []
    for w, period, q, v, s, r, note in weeks:
        daily_q = round(DAILY_TOTAL * q / 1000) * 1000
        daily_v = round(DAILY_TOTAL * v / 1000) * 1000
        daily_s = round(DAILY_TOTAL * s / 1000) * 1000
        daily_r = round(DAILY_TOTAL * r / 1000) * 1000
        daily_total = daily_q + daily_v + daily_s + daily_r
        weekly = daily_total * 7 if w != "Tuần 8" else daily_total * 10
        out.append((w, period, daily_q, daily_v, daily_s, daily_r, daily_total, weekly, note))
    return out

# =========================================================
# 7) KPI TARGETS
# =========================================================
def build_kpi(campaigns):
    rows = []
    for c in campaigns:
        funnel = c["Funnel"]
        level = c["Level"]
        # CPL targets (VND)
        cpl_map = {
            ("Quiz", "Tiểu học"): 80_000,  ("Quiz", "THCS"): 100_000, ("Quiz", "THPT"): 120_000,
            ("VSL",  "Tiểu học"): 60_000,  ("VSL",  "THCS"): 75_000,  ("VSL",  "THPT"): 90_000,
            ("Sales","Tiểu học"): 200_000, ("Sales","THCS"): 250_000, ("Sales","THPT"): 350_000,
        }
        cpl = cpl_map[(funnel, level)]
        # CPM, CTR, CPC
        rows.append({
            "Campaign": c["Campaign Name"],
            "Funnel": funnel,
            "Slot Goal": c["Slot Goal"],
            "Target CPM": "150,000 - 250,000",
            "Target CTR": "≥ 1.8%" if funnel == "Quiz" else "≥ 2.2%",
            "Target CPC": "≤ 8,000",
            "Target CPL (VND)": cpl,
            "Target Booking Rate": "20-35%",
            "Target Show Rate": "60-80%",
            "Target Close Rate": "15-30%" if funnel != "Sales" else "25-40%",
            "Min Daily": int(c["Daily Budget (VND)"] * 0.5),
            "Scale Threshold": f"CPL ≤ {int(cpl*0.8):,}đ + CTR ≥ 2.5% — tăng 20%/ngày",
            "Kill Threshold":  f"CPL > {int(cpl*1.5):,}đ sau 3 ngày",
        })
    return rows

# =========================================================
# 8) CHECKLIST GO-LIVE
# =========================================================
CHECKLIST = [
    ("Pixel/CAPI", "Cài Meta Pixel base trên tất cả LP"),
    ("Pixel/CAPI", "Setup CAPI server-side (Cloudflare Worker) cho 8 events"),
    ("Pixel/CAPI", "Test 8 events với Pixel Helper + Test Events"),
    ("Pixel/CAPI", "Map Lead_QuizComplete → Custom Conversion (priority 4)"),
    ("Pixel/CAPI", "Map Lead_FormSubmit → Custom Conversion (priority 6)"),
    ("Pixel/CAPI", "CAPI: Booking_Confirmed + Enroll_Paid từ GHL/Pancake CRM"),
    ("Audiences",  "Upload CSV LeadDB 12 tháng → CA_LeadDB_All"),
    ("Audiences",  "Tạo 5 LAL từ Custom Audiences"),
    ("Audiences",  "Cấu hình Saved Audiences theo Region"),
    ("Audiences",  "Setup Exclusions: Existing Leads + Customers"),
    ("Creative",   "Render 30 ads × 3 ratio (9:16 / 1:1 / 4:5) = 90 file"),
    ("Creative",   "Upload subtitle .srt cho mọi video"),
    ("Creative",   "Setup Instant Form (Lead Gen) cho fallback (no LP)"),
    ("Creative",   "A/B test 3 hook/ad set (Dynamic Creative)"),
    ("Campaign",   "Tạo 24 Campaigns + 96 Ad Sets từ template"),
    ("Campaign",   "Set ABO 7 ngày đầu → CBO sau Day 7"),
    ("Campaign",   "Đặt Special Ad Category = None (xác nhận không trigger)"),
    ("Campaign",   "Setup Naming Convention: REGION_LEVEL_FUNNEL_LionCamp_2026"),
    ("Tracking",   "UTM chuẩn: utm_source=fb&utm_campaign={campaign}&utm_content={ad}"),
    ("Tracking",   "Setup GA4 + UTM dashboard (Looker Studio)"),
    ("Tracking",   "GHL pipeline: Lead → Booked → Show → Enroll"),
    ("Tracking",   "Pancake CRM auto-tag theo source ads"),
    ("Tracking",   "Slack alert khi Booking_Confirmed (CAPI hook)"),
    ("Landing",    "Quiz LP: ≤ 4 câu hỏi + filter (Lớp/Khu vực/SĐT)"),
    ("Landing",    "VSL LP: video ≤ 90s + CTA Đặt lịch"),
    ("Landing",    "Sales LP: countdown + scarcity + 3 CTA buttons"),
    ("Landing",    "Mobile-first + LCP ≤ 2.5s"),
    ("Landing",    "Privacy Policy + Terms + Cookie banner"),
    ("Sales/Ops",  "Telesales script 3 calls (Confirm / Tư vấn / Close)"),
    ("Sales/Ops",  "Slot tracker realtime (Google Sheet liên kết LP scarcity)"),
    ("Sales/Ops",  "Lịch trực telesales 12-21h × 7 ngày/tuần"),
    ("Sales/Ops",  "SLA: Gọi lead trong ≤ 5 phút"),
    ("Budget",     "Set daily budget 24 campaigns"),
    ("Budget",     "Setup spend cap toàn account = 7tr/ngày + buffer 20%"),
    ("Budget",     "Cảnh báo email khi spend > 80% daily"),
    ("Reporting",  "Daily report 8h sáng (CPL/CTR/Lead/Booking)"),
    ("Reporting",  "Weekly review meeting Thứ 2 — Cut/Scale quyết định"),
    ("Reporting",  "Export báo cáo cuối campaign (Day 60)"),
    ("Compliance", "Đọc lại policy của Meta về quảng cáo giáo dục trẻ em"),
    ("Compliance", "Avoid claims tuyệt đối (best/100%/guarantee)"),
    ("Compliance", "Không upload ảnh trẻ em không có consent ký"),
    ("Compliance", "Có disclaimer “Kết quả tuỳ học sinh”"),
]

# =========================================================
# 9) AD COPY FULL (theme → bộ copy)
# =========================================================
AD_COPY_FULL = [
    ("Pain (TH)",        "Mùa hè 2 tháng đang đến gần. Con bạn sẽ làm gì? Ở nhà với điện thoại, máy tính bảng, TV — hay phát triển kỹ năng + tiếng Anh + sức khoẻ? Lion Camp — Trại hè bán trú tại Gò Vấp/Bình Tân, học qua dự án + thể thao + kỹ năng sống.", "Hè Vàng — Đừng để lãng phí", "Đăng ký giữ chỗ"),
    ("Outcome (TH)",     "Sau 4 tuần tại Lion Camp, con bạn sẽ: (1) Tự tin giao tiếp tiếng Anh cơ bản, (2) Có kỹ năng sống độc lập, (3) Vận động + sức khoẻ tốt, (4) Có 30+ bạn mới cùng độ tuổi.", "Lion Camp Tiểu học — 4 tuần", "Tìm hiểu chương trình"),
    ("Urgency (TH)",     "🔥 Còn 18/50 suất Tiểu học Gò Vấp — Khai giảng 15/06. Phụ huynh đăng ký trong tuần này được tặng bộ sách + áo trại trị giá 500k.", "Sắp hết suất — Giữ chỗ ngay", "Giữ chỗ ngay"),
    ("Pain (THCS)",      "Con bạn lớp 6-9 đang ở giai đoạn quan trọng nhất. Game, TikTok, lười học — nếu không can thiệp sớm sẽ ảnh hưởng cả tương lai. Lion Camp THCS giúp con tìm lại định hướng + động lực.", "THCS — Tuổi vàng định hướng", "Test định hướng MIỄN PHÍ"),
    ("Identity (THCS)",  "Độ tuổi 11-15 quyết định 80% tương lai — Đây là lúc não con phát triển nhanh nhất. 4 tuần Lion Camp = Năng lượng cho cả năm học sau.", "Khai phá tiềm năng THCS", "Xem chương trình"),
    ("Urgency (THCS)",   "🔥 Đã full 70% suất THCS Gò Vấp — Còn 30 suất cuối. Đăng ký sớm = Có lịch interview với học sinh + Tour campus miễn phí.", "Sắp đóng đăng ký THCS", "Giữ chỗ"),
    ("Pain (THPT)",      "Con bạn lớp 10-12 đã biết mình muốn học gì sau lớp 12 chưa? 70% học sinh chọn ngành theo cảm tính — và hối tiếc sau 1 năm Đại học. Trại hè Lion Camp THPT giúp con định hướng nghề nghiệp khoa học.", "THPT — Định hướng nghề nghiệp", "Test định hướng"),
    ("StudyAbroad (THPT)","Chuẩn bị du học từ THPT là lựa chọn của top 10% phụ huynh khôn ngoan. Lion Camp THPT có roadmap 3 năm: IELTS → SAT → Hồ sơ. Mentor là alumni NUS, NTU, RMIT.", "Du học roadmap 3 năm", "Tư vấn 1:1"),
    ("Urgency (THPT)",   "🔥 Đóng đăng ký THPT sau 7 ngày — Còn 22 suất. Mentor 1:1 chỉ dành cho 30 học sinh đăng ký sớm.", "Final Call — THPT", "Giữ chỗ ngay"),
    ("Local (Tỉnh)",     "Phụ huynh Cần Giuộc + Rạch Giá: Trại hè Tiểu học chất lượng Sài Gòn — Ngay tại địa phương. Đưa đón tận trường, học phí ưu đãi 25%.", "Trại hè Tiểu học gần nhà", "Tư vấn miễn phí"),
    ("Bonus",            "🎁 Đăng ký trước 30/04 — Tặng combo: Sách Tiếng Anh chuẩn Cambridge + Áo trại + Voucher 500k cho khoá tiếp theo.", "Early Bird — Tặng 500k", "Nhận ưu đãi"),
    ("Social Proof",     "1.200+ phụ huynh đã chọn Lion Camp 2 năm qua. Tỷ lệ tái đăng ký 78%. Đọc review thật từ phụ huynh tại fanpage.", "1.200+ phụ huynh tin tưởng", "Đọc reviews"),
    ("ROI Parent",       "Đầu tư 1 hè = Tiết kiệm 2 năm sai hướng. Phụ huynh thông minh đầu tư cho định hướng — không phải đợi đến lớp 12 mới lo.", "ROI rõ ràng cho phụ huynh", "Tư vấn ngay"),
    ("Schedule",         "Lịch trại hè 4 tuần (15/06 → 15/07): Sáng học chính khoá, chiều dự án + thể thao, tối kỹ năng. Cuối tuần được về nhà.", "Lịch trại hè 2026", "Xem lịch chi tiết"),
    ("Last Call",        "🚨 Còn đúng 14 suất THPT cuối — Đóng đăng ký 25/05. Sau 25/05 không nhận thêm dù có chỗ trống.", "FINAL CALL — Đóng 25/05", "Đăng ký gấp"),
]

# =========================================================
# 10) CREATIVE BRIEF
# =========================================================
CREATIVE_BRIEFS = [
    ("CB_TH_01",   "Video 9:16",  "1080x1920", "30s",  "Cảnh: Con ở nhà cầm điện thoại → cut → Con ở Lion Camp cười với bạn",   "VO Vietnamese + caption", "Pain → Outcome contrast"),
    ("CB_TH_02",   "Video 1:1",   "1080x1080", "60s",  "Một ngày của Lion Camp: dậy → học → thể thao → dự án → về",            "VO + sub",                "A day in the life"),
    ("CB_THCS_01", "Video 9:16",  "1080x1920", "45s",  "Phỏng vấn 5 học sinh THCS sau trại hè: 'Em thay đổi gì?'",              "Real voice + sub",        "Testimonial montage"),
    ("CB_THCS_02", "Carousel 1:1","1080x1080", "static","6 ảnh: hoạt động + outcome + reviews + pricing + CTA",                 "Headline cards",          "Carousel hook"),
    ("CB_THPT_01", "Video 9:16",  "1080x1920", "60s",  "Alumni LionCamp đỗ NUS/NTU chia sẻ: 'Lion Camp đã cho em định hướng'",   "Real voice",              "Authority + Aspiration"),
    ("CB_THPT_02", "Video 16:9",  "1920x1080", "90s",  "VSL hoàn chỉnh: Hook → Pain → Solution → Proof → Offer → Close",        "Pro VO + caption",        "Sales VSL"),
    ("CB_LOCAL_01","Video 9:16",  "1080x1920", "30s",  "Drone shot Cần Giuộc + Rạch Giá → Lion Camp campus local",              "VO local accent",         "Local connection"),
    ("CB_URGENCY_01","Static 1:1","1080x1080", "static","Countdown thật: 'Còn 18 suất' + Khai giảng 15/06",                     "Headline lớn",            "Scarcity static"),
    ("CB_BONUS_01","Video 1:1",   "1080x1080", "30s",  "Unboxing combo Early Bird: sách + áo + voucher",                        "Energetic music",         "Gift reveal"),
    ("CB_SCHEDULE_01","Carousel","1080x1080", "static","6 cards: Lịch tuần 1 → tuần 4 + lịch về nhà",                          "Numbered cards",          "Schedule clarity"),
]

# =========================================================
# WRITER: XLSX
# =========================================================
THIN  = Side(style="thin",  color="CCCCCC")
BORD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR   = PatternFill("solid", fgColor="1F4E78")
HDRF  = Font(bold=True, color="FFFFFF", name="Arial")
TITLE = Font(bold=True, size=14, color="FFFFFF", name="Arial")
TFILL = PatternFill("solid", fgColor="305496")
SUB   = PatternFill("solid", fgColor="D9E1F2")
BODY  = Font(name="Arial", size=10)

def style_sheet(ws, title, headers, data, widths=None):
    ws.cell(row=1, column=1, value=title).font = TITLE
    ws.cell(row=1, column=1).fill = TFILL
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    if len(headers) > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = HDRF
        c.fill = HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORD
    ws.row_dimensions[3].height = 32
    for r_idx, row in enumerate(data, start=4):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORD
            if (r_idx % 2) == 0:
                cell.fill = SUB
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

def fmt_money(v):
    if isinstance(v, (int, float)):
        return f"{int(v):,}"
    return v

def write_xlsx(campaigns, adsets, ads, pacing, kpi):
    wb = Workbook()
    wb.remove(wb.active)

    # 0_HƯỚNG_DẪN
    ws = wb.create_sheet("0_HƯỚNG_DẪN")
    style_sheet(ws,
        "📘 LION CAMP 2026 — META ADS BULK v2 (60 NGÀY CUỐI)",
        ["Topic", "Chi tiết"],
        [
            ["Mục tiêu",       "Fill 600 slots (GV+BT+CL+RG) trong 60 ngày — Khai giảng 15/06/2026"],
            ["Total Budget",   f"{TOTAL_BUDGET:,} VND ({DAILY_TOTAL:,}/day × {DAYS} ngày)"],
            ["Cấu trúc",       "24 Campaigns × 4 Ad Sets = 96 Ad Sets + 30 Ads"],
            ["Naming",         "[LOC]_[LEVEL]_[FUNNEL]_LionCamp_2026 (LOC: GV/BT/CL/RG, LEVEL: TH/THCS/THPT, FUNNEL: Quiz/VSL/Sales)"],
            ["Funnel split",   "Quiz 50% • VSL 30% • Sales 20%"],
            ["Level split",    "TH 35% • THCS 30% • THPT 35%"],
            ["Region weight",  "GV 50/50 (THPT,THCS) • TH split GV 30 / BT 30 / CL 20 / RG 20"],
            ["Adset types",    "Broad AI 30% • Interest 25% • LAL 20% • RMK 25%"],
            ["Pixel events",   "8 events có giá trị + CAPI server-side"],
            ["Audiences",      "≥17 audiences (10 CA + 4 LAL + 3 Saved)"],
            ["KPI",            "CPL 80-150k • Booking 20-35% • Show 60-80% • Close 15-30%"],
            ["Phase 1 (D1-30)","Scale Quiz, test creative, build pixel data"],
            ["Phase 2 (D30-50)","Tăng VSL + Sales, giảm Quiz"],
            ["Phase 3 (D50-60)","ALL-IN: 60% Sales + 40% Retarget — chốt slot cuối"],
            ["Tracking",       "GA4 + UTM + GHL pipeline + Pancake CRM"],
            ["Compliance",     "Special Ad Category: None (kiểm tra) • Có disclaimer • Không claim tuyệt đối"],
            ["Sheet hierarchy","1_Pixel → 2_Audiences → 3_Campaigns → 4_AdSets → 5_Ads → 6_Copy → 7_Brief → 8_Pacing → 9_KPI → 10_Checklist"],
        ],
        widths=[26, 110])

    # 1_Pixel_Events
    ws = wb.create_sheet("1_Pixel_Events")
    style_sheet(ws,
        "📊 META PIXEL + 8 CONVERSION EVENTS (CAPI ready)",
        ["Priority", "Event Name", "Trigger", "Value (VND)", "Custom Conversion?"],
        [list(r) for r in PIXEL_EVENTS],
        widths=[10, 28, 50, 18, 22])

    # 2_Audiences
    ws = wb.create_sheet("2_Audiences")
    style_sheet(ws,
        "🎯 AUDIENCE STRATEGY — 17 AUDIENCES",
        ["Type", "Audience Name", "Source/Definition", "Size Est.", "Duration", "Used In"],
        [list(r) for r in AUDIENCES],
        widths=[10, 28, 50, 14, 12, 28])

    # 3_Campaigns
    ws = wb.create_sheet("3_Campaigns")
    headers = ["Campaign Name","Region","Level","Funnel","Slot Goal","Objective","CBO?","Daily Budget (VND)","Total Budget (VND)","Bid Strategy","Start","End","Special Ad Category"]
    data = [[c["Campaign Name"], c["Region"], c["Level"], c["Funnel"], c["Slot Goal"], c["Objective"], c["CBO"], fmt_money(c["Daily Budget (VND)"]), fmt_money(c["Total Budget (VND)"]), c["Bid Strategy"], c["Start"], c["End"], c["Special Ad Category"]] for c in campaigns]
    style_sheet(ws,
        f"🎯 24 CAMPAIGNS — Lion Camp 2026 ({TOTAL_BUDGET:,} VND / {DAYS} ngày)",
        headers, data,
        widths=[34, 12, 10, 10, 9, 36, 26, 18, 18, 22, 12, 12, 18])

    # 4_AdSets
    ws = wb.create_sheet("4_AdSets")
    headers = ["Campaign","Ad Set Name","Audience","Exclude","Locations","Age","Gender","Placement","Daily Budget (VND)","Optimization","Attribution","Pacing"]
    data = [[a["Campaign"], a["Ad Set Name"], a["Audience"], a["Exclude"], a["Locations"], a["Age"], a["Gender"], a["Placement"], fmt_money(a["Daily Budget (VND)"]), a["Optimization"], a["Attribution"], a["Pacing"]] for a in adsets]
    style_sheet(ws,
        "📁 96 AD SETS (24 Campaigns × 4 Ad Sets)",
        headers, data,
        widths=[34, 38, 38, 28, 38, 8, 8, 28, 16, 30, 18, 12])

    # 5_Ads_Creative
    ws = wb.create_sheet("5_Ads_Creative")
    headers = ["Campaign","Ad Set","Ad Name","Format","Theme","Primary Text Hook","Headline","CTA Button","Landing URL"]
    data = [[a["Campaign"], a["Ad Set"], a["Ad Name"], a["Format"], a["Theme"], a["Primary Text Hook"], a["Headline"], a["CTA Button"], a["Landing URL"]] for a in ads]
    style_sheet(ws,
        f"🎨 {len(ads)} ADS — 30 unique creatives × adsets matched theo funnel",
        headers, data,
        widths=[34, 38, 42, 32, 14, 50, 32, 18, 60])

    # 6_Ad_Copy_Full
    ws = wb.create_sheet("6_Ad_Copy_Full")
    style_sheet(ws,
        "✍️ AD COPY FULL — Vietnamese (15 themes)",
        ["Theme", "Primary Text (Full)", "Headline (40 chars)", "Description / CTA"],
        [list(r) for r in AD_COPY_FULL],
        widths=[18, 80, 32, 26])

    # 7_Creative_Brief
    ws = wb.create_sheet("7_Creative_Brief")
    style_sheet(ws,
        "🎬 CREATIVE BRIEF — Cho team Design & Video (10 assets)",
        ["Asset ID", "Format", "Dimension", "Duration", "Scene/Script", "Voice-over/Caption", "Key Visual"],
        [list(r) for r in CREATIVE_BRIEFS],
        widths=[14, 14, 14, 10, 60, 26, 26])

    # 8_Budget_Pacing
    ws = wb.create_sheet("8_Budget_Pacing")
    headers = ["Tuần","Period","Quiz","VSL","Sales","RMK","Total Daily","Weekly Total","Note"]
    data = [[w, p, fmt_money(q), fmt_money(v), fmt_money(s), fmt_money(r), fmt_money(t), fmt_money(wk), n] for (w,p,q,v,s,r,t,wk,n) in pacing]
    style_sheet(ws,
        "💰 BUDGET PACING — 60 NGÀY (8 tuần)",
        headers, data,
        widths=[10, 16, 14, 14, 14, 14, 14, 16, 40])

    # 9_KPI_Targets
    ws = wb.create_sheet("9_KPI_Targets")
    headers = ["Campaign","Funnel","Slot Goal","Target CPM","Target CTR","Target CPC","Target CPL (VND)","Booking Rate","Show Rate","Close Rate","Min Daily","Scale Threshold","Kill Threshold"]
    data = [[k["Campaign"], k["Funnel"], k["Slot Goal"], k["Target CPM"], k["Target CTR"], k["Target CPC"], fmt_money(k["Target CPL (VND)"]), k["Target Booking Rate"], k["Target Show Rate"], k["Target Close Rate"], fmt_money(k["Min Daily"]), k["Scale Threshold"], k["Kill Threshold"]] for k in kpi]
    style_sheet(ws,
        "🎯 KPI TARGETS — Per Campaign (24 rows)",
        headers, data,
        widths=[34, 10, 9, 18, 12, 12, 18, 14, 14, 14, 14, 36, 28])

    # 10_Checklist_GoLive
    ws = wb.create_sheet("10_Checklist_GoLive")
    style_sheet(ws,
        "✅ META ADS CHECKLIST — Trước Go-Live",
        ["Nhóm", "Item", "Done"],
        [[g, item, ""] for (g, item) in CHECKLIST],
        widths=[14, 80, 8])

    out_path = os.path.join(OUT, "MetaAds_LionCamp_2026_v2.xlsx")
    wb.save(out_path)
    return out_path

# =========================================================
# WRITER: CSVs (for Apps Script)
# =========================================================
def write_csvs(campaigns, adsets, ads, pacing, kpi):
    def write(name, rows, headers):
        path = os.path.join(OUT, name)
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in rows:
                w.writerow(r)
        return path

    write("3_Campaigns.csv",
          [[c["Campaign Name"], c["Region"], c["Level"], c["Funnel"], c["Slot Goal"], c["Objective"], c["CBO"], c["Daily Budget (VND)"], c["Total Budget (VND)"], c["Bid Strategy"], c["Start"], c["End"], c["Special Ad Category"]] for c in campaigns],
          ["Campaign Name","Region","Level","Funnel","Slot Goal","Objective","CBO?","Daily Budget (VND)","Total Budget (VND)","Bid Strategy","Start","End","Special Ad Category"])

    write("4_AdSets.csv",
          [[a["Campaign"], a["Ad Set Name"], a["Audience"], a["Exclude"], a["Locations"], a["Age"], a["Gender"], a["Placement"], a["Daily Budget (VND)"], a["Optimization"], a["Attribution"], a["Pacing"]] for a in adsets],
          ["Campaign","Ad Set Name","Audience","Exclude","Locations","Age","Gender","Placement","Daily Budget (VND)","Optimization","Attribution","Pacing"])

    write("5_Ads_Creative.csv",
          [[a["Campaign"], a["Ad Set"], a["Ad Name"], a["Format"], a["Theme"], a["Primary Text Hook"], a["Headline"], a["CTA Button"], a["Landing URL"]] for a in ads],
          ["Campaign","Ad Set","Ad Name","Format","Theme","Primary Text Hook","Headline","CTA Button","Landing URL"])

    write("9_KPI_Targets.csv",
          [[k["Campaign"], k["Funnel"], k["Slot Goal"], k["Target CPM"], k["Target CTR"], k["Target CPC"], k["Target CPL (VND)"], k["Target Booking Rate"], k["Target Show Rate"], k["Target Close Rate"], k["Min Daily"], k["Scale Threshold"], k["Kill Threshold"]] for k in kpi],
          ["Campaign","Funnel","Slot Goal","Target CPM","Target CTR","Target CPC","Target CPL (VND)","Booking Rate","Show Rate","Close Rate","Min Daily","Scale Threshold","Kill Threshold"])

# =========================================================
# WRITER: JSON dump (for Apps Script payload)
# =========================================================
def write_json(campaigns, adsets, ads, pacing, kpi):
    payload = {
        "version": "v2",
        "total_budget_vnd": TOTAL_BUDGET,
        "days": DAYS,
        "daily_total_vnd": DAILY_TOTAL,
        "guide": [
            ["Mục tiêu",       "Fill 600 slots (GV+BT+CL+RG) trong 60 ngày — Khai giảng 15/06/2026"],
            ["Total Budget",   f"{TOTAL_BUDGET:,} VND ({DAILY_TOTAL:,}/day × {DAYS} ngày)"],
            ["Cấu trúc",       "24 Campaigns × 4 Ad Sets = 96 Ad Sets + 30 Ads (mapped)"],
            ["Naming",         "[LOC]_[LEVEL]_[FUNNEL]_LionCamp_2026"],
            ["Funnel split",   "Quiz 50% • VSL 30% • Sales 20%"],
            ["Level split",    "TH 35% • THCS 30% • THPT 35%"],
            ["Region weight",  "THPT/THCS = GV 50 / BT 50 — TH = GV 30 / BT 30 / CL 20 / RG 20"],
            ["Adset types",    "Broad AI 30% • Interest 25% • LAL 20% • RMK 25%"],
            ["Pixel events",   "8 events có giá trị + CAPI server-side"],
            ["Audiences",      "17 audiences (10 CA + 4 LAL + 3 Saved)"],
            ["KPI",            "CPL 80-150k • Booking 20-35% • Show 60-80% • Close 15-30%"],
            ["Phase 1 (D1-30)","Scale Quiz, test creative"],
            ["Phase 2 (D30-50)","Tăng VSL + Sales"],
            ["Phase 3 (D50-60)","ALL-IN closing 10 ngày cuối"],
        ],
        "pixel_events": [list(r) for r in PIXEL_EVENTS],
        "audiences":    [list(r) for r in AUDIENCES],
        "campaigns":    [
            [c["Campaign Name"], c["Region"], c["Level"], c["Funnel"], c["Slot Goal"], c["Objective"], c["CBO"], c["Daily Budget (VND)"], c["Total Budget (VND)"], c["Bid Strategy"], c["Start"], c["End"], c["Special Ad Category"]]
            for c in campaigns
        ],
        "adsets": [
            [a["Campaign"], a["Ad Set Name"], a["Audience"], a["Exclude"], a["Locations"], a["Age"], a["Gender"], a["Placement"], a["Daily Budget (VND)"], a["Optimization"], a["Attribution"], a["Pacing"]]
            for a in adsets
        ],
        "ads": [
            [a["Campaign"], a["Ad Set"], a["Ad Name"], a["Format"], a["Theme"], a["Primary Text Hook"], a["Headline"], a["CTA Button"], a["Landing URL"]]
            for a in ads
        ],
        "ad_copy_full":  [list(r) for r in AD_COPY_FULL],
        "creative_briefs": [list(r) for r in CREATIVE_BRIEFS],
        "pacing":  [list(r) for r in pacing],
        "kpi":     [
            [k["Campaign"], k["Funnel"], k["Slot Goal"], k["Target CPM"], k["Target CTR"], k["Target CPC"], k["Target CPL (VND)"], k["Target Booking Rate"], k["Target Show Rate"], k["Target Close Rate"], k["Min Daily"], k["Scale Threshold"], k["Kill Threshold"]]
            for k in kpi
        ],
        "checklist": [list(r) for r in CHECKLIST],
    }
    path = os.path.join(OUT, "v2_payload.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return path

# =========================================================
# MAIN
# =========================================================
if __name__ == "__main__":
    campaigns = build_campaigns()
    adsets = build_adsets(campaigns)
    ads = build_ads(adsets)
    pacing = build_pacing(campaigns)
    kpi = build_kpi(campaigns)

    print(f"Campaigns: {len(campaigns)}")
    print(f"Ad Sets:   {len(adsets)}")
    print(f"Ads:       {len(ads)}")
    print(f"Pacing:    {len(pacing)} weeks")
    print(f"KPI:       {len(kpi)}")

    xlsx_path = write_xlsx(campaigns, adsets, ads, pacing, kpi)
    write_csvs(campaigns, adsets, ads, pacing, kpi)
    json_path = write_json(campaigns, adsets, ads, pacing, kpi)

    print(f"\nXLSX → {xlsx_path}")
    print(f"JSON → {json_path}")
